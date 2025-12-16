from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import logout, login, authenticate
from django.contrib import messages
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Count, F, Q, Prefetch, ExpressionWrapper, fields
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.contrib.auth.models import User
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .forms import CustomUserCreationForm, CustomAuthenticationForm
from django.contrib.auth.forms import AuthenticationForm
from .models import Order, OrderItem, MenuItem, Group, RegistrationPin, Category, RoomBill
import json
import decimal


class DecimalEncoder(json.JSONEncoder):
    """Custom JSON encoder for Decimal objects"""
    def default(self, obj):
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        return super().default(obj)

def get_order_identifier(order):
    """
    Returns the order identifier combining room number and client name if available.
    If only room_number exists, shows that.
    If only client_identifier exists, shows that.
    If both exist, shows: "Habitación X - Cliente Y"
    """
    if order.room_number and order.client_identifier:
        return f"Habitación {order.room_number} - {order.client_identifier}"
    elif order.room_number:
        return f"Habitación {order.room_number}"
    else:
        return order.client_identifier

def home(request):
    if request.user.is_authenticated:
        if request.user.is_superuser or request.user.groups.filter(name='Administrador').exists():
            return redirect('restaurant:admin_dashboard')
        elif request.user.groups.filter(name='Recepcionista').exists():
            return redirect('restaurant:receptionist_dashboard')
        elif request.user.groups.filter(name='Cocinero').exists():
            return redirect('restaurant:cook_dashboard')
        elif request.user.groups.filter(name='Garzón').exists():
            return redirect('restaurant:waiter_dashboard')
        else:
            # No role assigned, render home
            return render(request, 'restaurant/home.html', {})
    else:
        return redirect('login')

def logout_view(request):
    logout(request)
    return redirect('login')

def signup_view(request):
    """
    Vista para el registro de nuevos usuarios.
    """
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Inicia sesión automáticamente para el nuevo usuario
            login(request, user)
            return redirect('restaurant:home')  # Redirige a home, que gestionará el dashboard correcto
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/signup.html', {'form': form})

def login_view(request):
    """
    Vista para el inicio de sesión de usuarios.
    """
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('restaurant:home')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'registration/login.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrador').exists())
def admin_dashboard(request):
    # ... (el resto de la vista se mantiene igual)
    import json
    from django.utils import timezone    

    # --- Optimización de Consultas ---
    # 1. Usar aggregate para obtener todos los contadores en una sola consulta
    today = timezone.localtime().date()
    stats = Order.objects.aggregate(
        total_today=Count('id', filter=Q(created_at__date=today)),
        preparing=Count('id', filter=Q(status='preparing', created_at__date=today)),
        ready=Count('id', filter=Q(status='ready', created_at__date=today)),
        completed=Count('id', filter=Q(status__in=['served', 'paid'], created_at__date=today)),
        total_sales_today=Sum('total_amount', filter=Q(status='paid', created_at__date=today))
    )

    # 2. Usar annotate para calcular el total de cada pedido en la DB (evita N+1)
    #    y prefetch_related para cargar los items eficientemente.
    # Ahora que total_amount se almacena directamente, no necesitamos annotate aquí.
    recent_orders = Order.objects.select_related('user').order_by('-created_at')[:10]

    # Nota: 'orders' aquí es Order.objects.all(), no recent_orders.
    # El resto de los objetos se cargan para el renderizado inicial o como fallback.
    # Optimización: select_related para evitar queries adicionales
    orders = Order.objects.select_related('user').all()
    menu_items = MenuItem.objects.all()
    groups = Group.objects.all()  # Obtener todos los grupos/roles
    user_role = request.user.groups.first().name if request.user.groups.exists() else None

    # JSON data for JavaScript
    orders_json = json.dumps([{
        'id': o.id,
        'identifier': get_order_identifier(o),
        'status': o.status,
        'user': o.user.username,
        'created_at': o.created_at.isoformat(),
        'total': float(o.total_amount), # Usar el nuevo campo total_amount
    } for o in orders], cls=DecimalEncoder)
    menu_items_json = json.dumps([{
        'id': m.id,
        'name': m.name,
        'description': m.description,
        'price': float(m.price),
        'category': m.category,
        'available': m.available,
        'image_url': m.image_url
    } for m in menu_items], cls=DecimalEncoder)
    return render(request, 'restaurant/admin_dashboard.html', {
        'orders': orders,
        'menu_items': menu_items,
        'groups': groups,  # Pasar los grupos a la plantilla
        'total_orders_today': stats.get('total_today', 0),
        'preparing_count': stats.get('preparing', 0),
        'ready_count': stats.get('ready', 0),
        'completed_count': stats.get('completed', 0),
        'total_sales_today': stats.get('total_sales_today', 0) or 0,
        'recent_orders': recent_orders,
        'user_role': user_role,
        'orders_json': orders_json,
        'menu_items_json': menu_items_json,
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Recepcionista').exists())
def receptionist_dashboard(request):
    from django.utils import timezone
    from django.db.models import Sum, F, Q

    # Unificamos los pedidos que necesitan acción de cobro: 'servido' y 'cargado a habitación'
    served_orders = Order.objects.filter(status__in=['served', 'charged_to_room']).order_by('-created_at')
    
    # Solo pedidos cargados a habitación que tengan room_number (no pedidos sin habitación)
    room_charge_orders = Order.objects.filter(
        status='charged_to_room',
        room_number__isnull=False,
        room_number__gt=''  # Excluir strings vacíos
    ).order_by('-created_at')



    today = timezone.now().date()
    total_sales_today = Order.objects.filter(
        status='paid',
        created_at__date=today
    ).aggregate(total=Sum('total_amount'))['total'] or 0 # Usar el nuevo campo total_amount

    user_role = request.user.groups.first().name if request.user.groups.exists() else None

    return render(request, 'restaurant/receptionist_dashboard.html', {
        'user_role': user_role,
        'served_orders': served_orders,
        'room_charge_orders': room_charge_orders,
        'total_sales_today': total_sales_today,
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Cocinero').exists())
def cook_dashboard(request):
    import json
    from django.utils import timezone

    try:
        # Consulta optimizada para obtener pedidos y sus items
        # Ahora que total_amount se almacena directamente, no necesitamos annotate aquí.
        orders = Order.objects.filter(
            status__in=['pending', 'preparing']
        ).select_related('user').prefetch_related(
            'orderitem_set__menu_item'
        ).order_by('created_at')

        # Preparar datos JSON para el frontend
        orders_data = []
        for order in orders:
            items_list = []
            for item in order.orderitem_set.all():
                if item.menu_item:  # Validar que menu_item no sea nulo
                    items_list.append({
                        'name': item.menu_item.name,
                        'quantity': item.quantity,
                        'note': item.note or '',
                        'is_prepared': getattr(item, 'is_prepared', False)  # Usar getattr para manejo seguro
                    })
            
            orders_data.append({
                'id': order.id,
                'status': order.status,
                'identifier': get_order_identifier(order),
                'user_username': order.user.username,
                'created_at': order.created_at.isoformat(),
                'items': items_list
            })

        return render(request, 'restaurant/cook_dashboard.html', {
            'orders': orders,
            'orders_json': json.dumps(orders_data, cls=DecimalEncoder),
            'PUSHER_KEY': settings.PUSHER_KEY,
            'PUSHER_CLUSTER': settings.PUSHER_CLUSTER,
        })
    except Exception as e:
        import traceback
        print(f"Error en cook_dashboard: {str(e)}")
        print(traceback.format_exc())
        raise

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Garzón').exists())
def waiter_dashboard(request):
    import json
    from django.utils import timezone
    from django.conf import settings

    try:
        # --- Datos para la toma de pedidos ---
        menu_items = MenuItem.objects.filter(available=True)
        # Obtenemos todas las categorías, las normalizamos en Python y eliminamos duplicados.
        # Esto asegura que 'Pizzas' y 'pizzas' se traten como una sola categoría en los filtros.
        all_categories = menu_items.values_list('category', flat=True)
        normalized_categories = {cat.capitalize() for cat in all_categories}
        categories = sorted(list(normalized_categories))
        menu_items_json = json.dumps([{
            'id': item.id, 
            'name': item.name, 
            'price': float(item.price),
            'image_url': item.image_url,
            'description': item.description
        } for item in menu_items], cls=DecimalEncoder)

        # --- Datos para el monitor de pedidos (carga inicial) ---
        # Usar only() para evitar campos que pueden no existir en la BD (como is_prepared)
        orders_for_monitor = Order.objects.filter(
            status__in=['pending', 'preparing', 'ready', 'served']
        ).select_related('user').prefetch_related(
            'orderitem_set__menu_item'
        ).order_by('created_at')

        initial_orders_data = []
        for order in orders_for_monitor:
            items_list = []
            for item in order.orderitem_set.all():
                if item.menu_item:  # Validar que menu_item no sea nulo
                    items_list.append({
                        'name': item.menu_item.name,
                        'quantity': item.quantity,
                        'is_prepared': item.is_prepared,
                    })
            
            initial_orders_data.append({
                'id': order.id,
                'status': order.status,
                'status_display': order.get_status_display(),
                'status_class': order.status_class,
                'identifier': get_order_identifier(order),
                'total': float(order.total_amount),
                'items': items_list
            })

        initial_orders_json = json.dumps(initial_orders_data, cls=DecimalEncoder)

        user_role = request.user.groups.first().name if request.user.groups.exists() else None

        return render(request, 'restaurant/waiter_dashboard.html', {
            'menu_items': menu_items,
            'categories': categories,
            'menu_items_json': menu_items_json,
            'initial_orders_json': initial_orders_json,
            'user_role': user_role,
            'pusher_key': settings.PUSHER_KEY,
            'pusher_cluster': settings.PUSHER_CLUSTER,
        })
    except Exception as e:
        import traceback
        print(f"Error en waiter_dashboard: {str(e)}")
        print(traceback.format_exc())
        raise

def public_menu_view(request):
    """
    Vista pública para mostrar el menú del restaurante, ideal para un código QR.
    No requiere autenticación.
    """
    # Obtener TODOS los items (disponibles e indisponibles) y ordenarlos por categoría
    menu_url = request.build_absolute_uri()
    menu_items = MenuItem.objects.all().order_by('category', 'name')

    # Agrupar items por categoría en un diccionario
    grouped_menu = {}
    for item in menu_items:
        category = item.category
        if not category: # Agrupar items sin categoría en 'Otros'
            category = 'Otros'
        if category not in grouped_menu:
            grouped_menu[category] = []
        grouped_menu[category].append(item)

    return render(request, 'restaurant/public_menu.html', {
        'grouped_menu': grouped_menu,
        # Puedes añadir más contexto si lo necesitas, como el nombre del restaurante
        'restaurant_name': 'Restaurante AbbaHotel',
        'menu_url': menu_url
    })

# Helper function to calculate subtotal for an order instance
def calculate_order_subtotal(order_instance):
    return order_instance.orderitem_set.aggregate(
        subtotal=Sum(F('menu_item__price') * F('quantity'), output_field=fields.DecimalField())
    )['subtotal'] or decimal.Decimal('0.00')
@csrf_exempt
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Garzón').exists())
def save_order(request):
    if request.method == 'POST':
        try:
            # Verificar que el usuario está autenticado
            if not request.user.is_authenticated:
                return JsonResponse({'success': False, 'error': 'User not authenticated'}, status=401)
            
            # Parsear los datos JSON
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError as e:
                print(f"❌ JSON decode error: {str(e)}, body: {request.body}")
                return JsonResponse({'success': False, 'error': f'Invalid JSON: {str(e)}'}, status=400)
            
            items = data.get('items', [])
            tip_amount = decimal.Decimal(str(data.get('tip_amount', '0.00')))

            # Validar que hay items
            if not items:
                return JsonResponse({'success': False, 'error': 'Order must have at least one item'}, status=400)

            # Protección contra doble envío: generar un hash basado en los items
            # para detectar si es un duplicado
            import hashlib
            items_hash = hashlib.md5(
                json.dumps(items, sort_keys=True).encode()
            ).hexdigest()
            
            # Verificar si existe una orden reciente con los mismos items
            # (creada en los últimos 5 segundos por el mismo usuario)
            from django.utils import timezone
            recent_time = timezone.now() - timezone.timedelta(seconds=5)
            
            # Buscar órdenes duplicadas recientes
            duplicate_orders = Order.objects.filter(
                user=request.user,
                created_at__gte=recent_time,
                client_identifier=data.get('client_identifier', 'Sin identificar'),
                room_number=data.get('room_number', '')
            )
            
            if duplicate_orders.exists():
                # Verificar si los items coinciden exactamente
                for dup_order in duplicate_orders:
                    dup_items = dup_order.orderitem_set.all()
                    if len(items) == dup_items.count():
                        # Verificar que los items coincidan
                        all_match = True
                        for item_data in items:
                            matching_item = dup_items.filter(
                                menu_item_id=item_data['id'],
                                quantity=int(item_data.get('quantity', 1))
                            ).exists()
                            if not matching_item:
                                all_match = False
                                break
                        
                        if all_match:
                            print(f"⚠️ Orden duplicada detectada. Retornando orden existente: {dup_order.id}")
                            # Retornar la orden existente para evitar duplicado
                            return JsonResponse({
                                'success': True,
                                'order_id': dup_order.id,
                                'message': 'Order already exists',
                                'is_duplicate': True,
                                'order': {
                                    'id': dup_order.id,
                                    'status': dup_order.status,
                                    'client_identifier': dup_order.client_identifier,
                                    'room_number': dup_order.room_number,
                                    'total_amount': float(dup_order.total_amount),
                                    'created_at': dup_order.created_at.isoformat(),
                                }
                            }, status=200)

            order = None
            # Usar una transacción atómica para asegurar la integridad de los datos.
            # O todo se crea, o nada se crea si hay un error.
            with transaction.atomic():
                subtotal = decimal.Decimal('0.00')
                
                # Pre-calculate subtotal before creating the order
                for item in items:
                    try:
                        menu_item = MenuItem.objects.get(id=item['id'])
                        subtotal += menu_item.price * int(item.get('quantity', 1))
                    except MenuItem.DoesNotExist as e:
                        print(f"❌ MenuItem not found: {item['id']}")
                        raise ValueError(f"MenuItem with id {item['id']} not found")
                    except KeyError as e:
                        print(f"❌ Missing key in item: {str(e)}")
                        raise ValueError(f"Missing required field in item: {str(e)}")
                
                # Create order with final total amount in one go
                try:
                    order = Order.objects.create(
                        client_identifier=data.get('client_identifier', 'Sin identificar'),
                        room_number=data.get('room_number', ''),
                        user=request.user, 
                        status='pending',
                        tip_amount=tip_amount,
                        total_amount=subtotal + tip_amount  # Set total_amount on creation
                    )
                    print(f"✅ Order created: {order.id}")
                except Exception as e:
                    print(f"❌ Error creating order: {str(e)}")
                    raise

                # Create order items
                for item in items:
                    try:
                        menu_item = MenuItem.objects.get(id=item['id'])
                        OrderItem.objects.create(
                            order=order,
                            menu_item=menu_item,
                            quantity=int(item.get('quantity', 1)),
                            note=str(item.get('note', ''))
                        )
                    except Exception as e:
                        print(f"❌ Error creating order item: {str(e)}")
                        raise
            
            if order:
                # Optimización: prefetch_related para cargar los items y sus menu_items de una vez
                order = Order.objects.prefetch_related('orderitem_set__menu_item').get(pk=order.id)
                items_list = order.orderitem_set.all()
                items_data = [{
                    'id': item.menu_item.id,
                    'name': item.menu_item.name,
                    'price': float(item.menu_item.price),
                    'quantity': item.quantity,
                    'note': item.note,
                } for item in items_list]
                
                # Mapeo de estados para mostrar en UI
                status_display_map = {
                    'pending': 'Pendiente',
                    'preparing': 'En Preparación',
                    'ready': 'Listo',
                    'served': 'Servido',
                    'paid': 'Pagado',
                    'cancelled': 'Cancelado',
                    'charged_to_room': 'Cargado a Habitación'
                }
                
                status_class_map = {
                    'pending': 'bg-yellow-100 text-yellow-800',
                    'preparing': 'bg-blue-100 text-blue-800',
                    'ready': 'bg-green-100 text-green-800',
                    'served': 'bg-purple-100 text-purple-800',
                    'paid': 'bg-green-600 text-white',
                    'cancelled': 'bg-red-100 text-red-800',
                    'charged_to_room': 'bg-indigo-100 text-indigo-800'
                }
                
                order_data = {
                    'success': True,
                    'order_id': order.id,
                    'order': {
                        'id': order.id,
                        'identifier': get_order_identifier(order),
                        'room_number': order.room_number,
                        'client_identifier': order.client_identifier,
                        'status': order.status,
                        'status_display': status_display_map.get(order.status, order.status),
                        'status_class': status_class_map.get(order.status, ''),
                        'items': items_data,
                        'total': float(order.total_amount),
                        'tip_amount': float(order.tip_amount),
                        'created_at': order.created_at.isoformat(),
                    }
                }
                print(f"✅ Order response sent successfully")
                return JsonResponse(order_data, safe=False)
            else:
                return JsonResponse({'success': False, 'error': 'Order could not be created'}, status=400)
                
        except MenuItem.DoesNotExist as e:
            print(f"❌ MenuItem.DoesNotExist: {str(e)}")
            return JsonResponse({'success': False, 'error': f'MenuItem not found'}, status=400)
        except ValueError as e:
            print(f"❌ ValueError: {str(e)}")
            return JsonResponse({'success': False, 'error': f'Validation error: {str(e)}'}, status=400)
        except Exception as e:
            import traceback
            error_msg = str(e)
            traceback_msg = traceback.format_exc()
            print(f"❌ Error in save_order: {error_msg}")
            print(f"Traceback:\n{traceback_msg}")
            return JsonResponse({'success': False, 'error': f'Server error: {error_msg}', 'detail': traceback_msg if settings.DEBUG else None}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method. Use POST.'}, status=405)

@csrf_exempt
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Garzón').exists())
def api_waiter_order_detail(request, pk):
    """
    API endpoint for waiters to get and update a specific order.
    GET: Returns the items of an order.
    PUT: Updates the items of an order.
    Optimized with prefetch_related to avoid N+1 queries.
    """
    try:
        # Optimización: prefetch_related para cargar los items y sus menu_items de una vez
        order = Order.objects.prefetch_related('orderitem_set__menu_item').get(pk=pk)
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Order not found'}, status=404)

    if request.method == 'GET':
        items = order.orderitem_set.all()
        items_data = [{ 
            'id': item.menu_item.id,
            'name': item.menu_item.name,
            'price': float(item.menu_item.price),
            'quantity': item.quantity,
            'note': item.note,
            'is_prepared': item.is_prepared,
        } for item in items]

        subtotal = sum(item['price'] * item['quantity'] for item in items_data)

        data = {
            'status': order.status,
            'items': items_data,
            'subtotal': subtotal, # Se agrega el subtotal para que el modal de pago funcione
            'total': subtotal + float(order.tip_amount), # El total ahora incluye la propina
            'identifier': get_order_identifier(order),
            'room_number': order.room_number, # Asegurarse de que este campo siempre esté presente
            'client_identifier': order.client_identifier,
        }
        return JsonResponse(data)

    if request.method == 'PUT':
        data = json.loads(request.body)
        items_data = data.get('items', [])

        with transaction.atomic():
            # Get existing items BEFORE any updates
            existing_items = list(order.orderitem_set.all())
            was_ready = order.status == 'ready'
            
            subtotal = decimal.Decimal('0.00')
            
            # For each item in the request, try to match with an existing item of the same type
            # If found, update it; otherwise, create a new one
            matched_existing_items = set()
            
            for item_data in items_data:
                menu_item_id = item_data['id']
                quantity = item_data['quantity']
                note = item_data.get('note', '')
                menu_item = MenuItem.objects.get(id=menu_item_id)
                
                # Try to find an unmatched existing item of the same type
                found_match = False
                for existing_item in existing_items:
                    if (existing_item.id not in matched_existing_items and 
                        existing_item.menu_item_id == menu_item_id):
                        # Update this item
                        existing_item.quantity = quantity
                        existing_item.note = note
                        # If order was "ready", mark this item as prepared
                        if was_ready:
                            existing_item.is_prepared = True
                        existing_item.save()
                        matched_existing_items.add(existing_item.id)
                        found_match = True
                        break
                
                if not found_match:
                    # Create new item (is_prepared=False for new items)
                    OrderItem.objects.create(
                        order=order,
                        menu_item=menu_item,
                        quantity=quantity,
                        note=note,
                        is_prepared=False
                    )
                
                subtotal += menu_item.price * quantity
            
            # Delete items that were not matched (i.e., removed from the order)
            for existing_item in existing_items:
                if existing_item.id not in matched_existing_items:
                    existing_item.delete()
            
            # If order was ready, change status to preparing
            if was_ready:
                order.status = 'preparing'
            
            order.total_amount = subtotal + order.tip_amount
            order.save(update_fields=['total_amount', 'status'])
            return JsonResponse({'success': True, 'order_id': order.id})

@csrf_exempt
@user_passes_test(lambda u: u.groups.filter(name='Recepcionista').exists())
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Cocinero').exists())
def update_order_status(request, order_id):
    if request.method == 'POST':
        status = request.POST.get('status')
        order = Order.objects.get(id=order_id)
        order.status = status
        order.save()
        return redirect('cook_dashboard')
    return redirect('cook_dashboard')

# --- API Views for Admin Dashboard ---

@csrf_exempt
@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Administrador', 'Cocinero']).exists())
@login_required
def api_orders(request):
    """
    GET: Returns a list of all orders with items (para cook dashboard y polling).
    POST: Creates a new order.
    """
    if request.method == 'GET':
        try:
            # Para cocina: solo órdenes pendientes y en preparación
            orders = Order.objects.filter(
                status__in=['pending', 'preparing']
            ).select_related('user').prefetch_related('orderitem_set__menu_item').order_by('created_at')
            
            data = []
            for o in orders:
                data.append({
                    'id': o.id,
                    'identifier': get_order_identifier(o),
                    'status': o.status,
                    'created_at': o.created_at.isoformat(),
                    'items': [{
                        'name': item.menu_item.name,
                        'quantity': item.quantity,
                        'note': item.note or '',
                        'is_prepared': item.is_prepared,
                    } for item in o.orderitem_set.all()]
                })
            return JsonResponse(data, safe=False)
        except Exception as e:
            print(f"[ERROR] api_orders GET: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order = Order.objects.create(
                client_identifier=data['client_identifier'],
                room_number=data.get('room_number'),
                user=request.user,
                status=data.get('status', 'pending')
            )
            
            return JsonResponse({
                'id': order.id,
                'identifier': get_order_identifier(order),
                'status': order.status
            }, status=201)
        except Exception as e:
            print(f"[ERROR] api_orders POST: {str(e)}")
            return JsonResponse({'error': str(e)}, status=400)



@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrador').exists())
def api_kitchen_orders(request):
    if request.method == 'GET':
        orders = Order.objects.filter(status__in=['pending', 'preparing'])
        data = [{
            'id': o.id,
            'identifier': get_order_identifier(o),
            'status': o.get_status_display(),
        } for o in orders]
        return JsonResponse(data, safe=False)

@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name__in=['Administrador', 'Recepcionista']).exists())
def api_order_detail(request, pk):
    """
    API endpoint for admin to get and update a specific order.
    GET: Returns the details of an order.
    PUT: Updates the order details.
    DELETE: Deletes the order.
    """
    try:
        order = Order.objects.get(pk=pk)
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Order not found'}, status=404)

    if request.method == 'GET':
        items = order.orderitem_set.all()
        items_data = [{
            'id': item.menu_item.id,
            'name': item.menu_item.name,
            'price': float(item.menu_item.price),
            'quantity': item.quantity,
            'note': item.note,
        } for item in items]

        subtotal = sum(item['price'] * item['quantity'] for item in items_data)
        tip_amount = float(order.tip_amount)

        data = {
            'id': order.id,
            'identifier': get_order_identifier(order),
            'status': order.status,
            'status_display': order.get_status_display(),
            'created_at': order.created_at.isoformat(),
            'user': order.user.username,
            'items': items_data,
            'subtotal': subtotal,
            'tip_amount': tip_amount,
            'total': subtotal + tip_amount,
        }
        return JsonResponse(data)

    if request.method == 'PUT':
        data = json.loads(request.body)
        order.client_identifier = data.get('client_identifier', order.client_identifier)
        order.room_number = data.get('room_number', order.room_number)
        order.status = data.get('status', order.status)

        order.save()
        return JsonResponse({'success': True, 'order_id': order.id})

    if request.method == 'DELETE':
        order.delete()
        return JsonResponse({'success': True}, status=204)

    return JsonResponse({'error': 'Invalid method'}, status=405)

@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name__in=['Administrador', 'Garzón', 'Cocinero', 'Recepcionista']).exists())
def api_order_status(request, pk):
    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
            order = Order.objects.get(pk=pk)
            new_status = data.get('status')
            
            if not new_status:
                return JsonResponse({'success': False, 'error': 'Status is required'}, status=400)
            
            order.status = new_status
            update_fields = ['status']
            
            # Solo recalcular el total si se envía explícitamente una nueva propina.
            # Esto evita que el modal de recepción recalcule el total con propinas antiguas.
            if new_status in ['paid', 'charged_to_room'] and 'tip_amount' in data:
                try:
                    order.tip_amount = decimal.Decimal(str(data['tip_amount'])) # Ensure it's a Decimal
                    subtotal = calculate_order_subtotal(order)
                    order.total_amount = subtotal + order.tip_amount
                    update_fields.extend(['tip_amount', 'total_amount'])
                except (ValueError, decimal.InvalidOperation) as e:
                    return JsonResponse({'success': False, 'error': f'Invalid tip amount: {str(e)}'}, status=400)
            
            # Handle payment method and reference
            if new_status in ['paid', 'charged_to_room']:
                if 'payment_method' in data:
                    order.payment_method = data['payment_method']
                    update_fields.append('payment_method')
                if 'payment_reference' in data and data['payment_reference']:
                    order.payment_reference = data['payment_reference']
                    update_fields.append('payment_reference')

            order.save(update_fields=update_fields)  # Only trigger signal for fields we actually changed
            return JsonResponse({'success': True, 'status': order.status, 'total_amount': float(order.total_amount)})
        except Order.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Order not found'}, status=404)
        except json.JSONDecodeError as e:
            return JsonResponse({'success': False, 'error': f'Invalid JSON: {str(e)}'}, status=400)
        except Exception as e:
            import traceback
            print(f"Error in api_order_status: {str(e)}")
            print(traceback.format_exc())
            return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'}, status=500)
    return JsonResponse({'error': 'Invalid method'}, status=405)

@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrador').exists())
def api_menu_items(request):
    """
    GET: Returns a list of all menu items.
    POST: Creates a new menu item.
    """
    if request.method == 'GET':
        menu_items = MenuItem.objects.all().order_by('category', 'name')
        data = [{
            'id': m.id, 'name': m.name, 'description': m.description,
            'price': float(m.price), 'category': m.category, 'available': m.available,
            'image_url': m.image_url
        } for m in menu_items]
        return JsonResponse(data, safe=False)

    if request.method == 'POST':
        try:
            # Parse JSON data
            data = json.loads(request.body)
            print(f"[DEBUG] POST data: {data}")
            
            # Parse available boolean
            available = data.get('available', True)
            if isinstance(available, str):
                available = available.lower() in ('true', '1', 'yes', 'on')
            
            # Validate required fields
            if 'name' not in data or not data['name'].strip():
                return JsonResponse({'error': 'Invalid data: name is required'}, status=400)
            if 'price' not in data:
                return JsonResponse({'error': 'Invalid data: price is required'}, status=400)
            
            # Convert price to Decimal to avoid comparison errors
            try:
                from decimal import Decimal
                price = Decimal(str(data['price']))
            except (ValueError, decimal.InvalidOperation):
                return JsonResponse({'error': 'Invalid data: price must be a valid number'}, status=400)
            
            item = MenuItem.objects.create(
                name=data['name'].strip(),
                description=data.get('description', ''),
                price=price,
                category=data.get('category', 'General'),
                available=available
            )
            
            return JsonResponse({
                'id': item.id, 'name': item.name, 'description': item.description,
                'price': float(item.price), 'category': item.category, 'available': item.available,
                'image_url': item.image_url
            }, status=201)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid data: malformed JSON'}, status=400)
        except ValueError as e:
            return JsonResponse({'error': f'Invalid data: {str(e)}'}, status=400)
        except Exception as e:
            print(f"[ERROR] Exception in POST: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': f'Invalid data: {str(e)}'}, status=400)

@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrador').exists())
def api_menu_item_detail(request, pk):
    """
    GET: Returns a single menu item.
    PUT: Updates a menu item completely.
    PATCH: Partially updates a menu item.
    DELETE: Deletes a menu item.
    """
    try:
        item = MenuItem.objects.get(pk=pk)
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Menu item not found'}, status=404)

    if request.method == 'GET':
        return JsonResponse({
            'id': item.id, 'name': item.name, 'description': item.description,
            'price': float(item.price), 'category': item.category, 'available': item.available,
            'image_url': item.image_url
        })

    if request.method in ['PUT', 'PATCH']:
        try:
            # Handle both FormData and JSON
            if request.content_type and 'multipart/form-data' in request.content_type:
                # FormData with possible file upload
                data = request.POST.dict()
            else:
                data = json.loads(request.body)
            print(f"[DEBUG] PATCH/PUT request for item {pk}: {data}")
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        
        def parse_available(val):
            """Helper function to parse available boolean from various formats"""
            if isinstance(val, bool):
                return val
            if isinstance(val, str):
                return val.lower() in ('true', '1', 'yes', 'on')
            return bool(val)
        
        def parse_price(val):
            """Helper function to parse price to Decimal"""
            if val is None:
                return None
            try:
                from decimal import Decimal
                return Decimal(str(val))
            except (ValueError, decimal.InvalidOperation):
                raise ValueError("Price must be a valid number")
            
        if request.method == 'PUT':
            # PUT: replace all fields
            item.name = data.get('name', item.name)
            item.description = data.get('description', item.description)
            if 'price' in data:
                item.price = parse_price(data['price'])
            item.category = data.get('category', item.category)
            if 'available' in data:
                item.available = parse_available(data['available'])
        else:
            # PATCH: only update provided fields
            if 'name' in data:
                item.name = data['name']
            if 'description' in data:
                item.description = data['description']
            if 'price' in data:
                item.price = parse_price(data['price'])
            if 'category' in data:
                item.category = data['category']
            if 'available' in data:
                item.available = parse_available(data['available'])
        
        # Handle image upload
        if 'image' in request.FILES:
            item.image = request.FILES['image']
        
        item.save()
        print(f"[DEBUG] Item {pk} saved. Available: {item.available}")
        return JsonResponse({
            'id': item.id, 'name': item.name, 'description': item.description,
            'price': float(item.price), 'category': item.category, 'available': item.available,
            'image_url': item.image_url
        })

    if request.method == 'DELETE':
        item.delete()
        return JsonResponse({'success': True}, status=204)

@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrador').exists())
def api_menu_item_upload_image(request, pk):
    """
    Upload image for a menu item directly to Cloudinary.
    Only handles POST requests with file upload.
    Returns the Cloudinary URL directly sin guardar en la BD.
    """
    import cloudinary.uploader
    
    try:
        item = MenuItem.objects.get(pk=pk)
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Menu item not found'}, status=404)
    
    if request.method == 'POST':
        if 'image' not in request.FILES:
            return JsonResponse({'error': 'No image file provided'}, status=400)
        
        try:
            image_file = request.FILES['image']
            print(f"[DEBUG] Uploading image for item {pk}: {image_file.name}")
            
            # Upload directly to Cloudinary using the file content
            result = cloudinary.uploader.upload(
                image_file.read(),
                folder=f'restaurant/menu_items',
                public_id=f'item_{pk}_{image_file.name.split(".")[0]}',
                overwrite=True,
                resource_type='auto'
            )
            
            cloudinary_url = result.get('secure_url')
            print(f"[DEBUG] Image uploaded to Cloudinary: {cloudinary_url}")
            
            # Store the Cloudinary URL public_id in the database (not the full URL)
            # This avoids Django treating it as a relative path
            item.image = f'cloudinary:{result.get("public_id")}'
            item.save()
            
            return JsonResponse({
                'id': item.id,
                'image_url': cloudinary_url,
                'message': 'Image uploaded successfully'
            }, status=200)
        except Exception as e:
            print(f"[ERROR] Image upload failed: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': f'Failed to upload image: {str(e)}'}, status=400)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)

@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrador').exists())
def api_menu_item_delete_image(request, pk):
    """
    Delete image for a menu item.
    Only handles POST requests.
    """
    try:
        item = MenuItem.objects.get(pk=pk)
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Menu item not found'}, status=404)
    
    if request.method == 'POST':
        try:
            if item.image:
                # Delete the file from storage
                item.image.delete()
                # Clear the image field
                item.image = None
                item.save()
                print(f"[DEBUG] Image deleted for item {pk}")
            
            return JsonResponse({
                'id': item.id,
                'message': 'Image deleted successfully'
            }, status=200)
        except Exception as e:
            print(f"[ERROR] Image deletion failed: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': f'Failed to delete image: {str(e)}'}, status=400)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Administrador', 'Recepcionista']).exists())
def api_orders_report(request):
    """
    API endpoint to get a filtered list of orders for reporting purposes.
    Optimized with select_related to avoid N+1 queries.
    """
    if request.method == 'GET':
        from django.db.models import Q
        from datetime import datetime, timedelta, date
        from django.utils import timezone
        
        # Optimización: select_related para el usuario (relación 1-to-N)
        orders = Order.objects.select_related('user').all()

        # Filtering
        search_query = request.GET.get('search', '')
        if search_query:
            # Búsqueda segura: Intenta buscar por ID si es un número, si no, solo por número de mesa.
            # Esto evita el error si el `search_query` no es un número válido.
            id_query = Q()
            if search_query.isdigit():
                id_query = Q(id=int(search_query))
            orders = orders.filter(id_query | Q(client_identifier__icontains=search_query) | Q(room_number__icontains=search_query))

        status_query = request.GET.get('status', '')
        if status_query:
            orders = orders.filter(status=status_query)

        date_from_query = request.GET.get('date_from', '')
        if date_from_query:
            try:
                date_from = datetime.strptime(date_from_query, '%Y-%m-%d').date()
                # Convert to datetime at start of day in current timezone, then to UTC for comparison
                dt_from = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
                orders = orders.filter(created_at__gte=dt_from)
            except ValueError:
                pass

        date_to_query = request.GET.get('date_to', '')
        if date_to_query:
            try:
                date_to = datetime.strptime(date_to_query, '%Y-%m-%d').date()
                # Convert to datetime at end of day in current timezone, then to UTC for comparison
                dt_to = timezone.make_aware(datetime.combine(date_to + timedelta(days=1), datetime.min.time()))
                orders = orders.filter(created_at__lt=dt_to)
            except ValueError:
                pass

        orders = orders.order_by('-created_at')

        # Prepare data for JSON response
        data = []
        for order in orders:
            data.append({
                'id': order.id,
                'identifier': get_order_identifier(order),
                'status': order.status,
                'status_display': order.get_status_display(),
                'status_class': order.status_class,
                'payment_method': order.payment_method,
                'payment_method_display': order.get_payment_method_display(),
                'created_at': order.created_at.isoformat(), # Ensure datetime is serializable
                'total': float(order.total_amount or 0), # Usar el nuevo campo total_amount
            })

        return JsonResponse(data, safe=False)

    return JsonResponse({'error': 'Invalid method'}, status=405)

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Administrador', 'Recepcionista']).exists())
def api_dashboard_charts(request):
    from django.db.models import Sum, Count, F, Q, Case, When, Value, IntegerField
    from django.db.models.functions import TruncDate, ExtractHour
    from django.utils import timezone
    from datetime import timedelta

    chart_type = request.GET.get('chart')
    today = timezone.localtime().date()

    if chart_type == 'sales_by_day':
        # Ventas de los últimos 7 días para pedidos pagados
        seven_days_ago = timezone.localtime().date() - timedelta(days=6)
        sales_data = Order.objects.filter(
            status='paid',
            created_at__date__gte=seven_days_ago
        ).annotate(
            date=TruncDate('created_at')
        ).values('date').annotate( # Use the new total_amount field
            daily_total=Sum('total_amount')
        ).order_by('date')

        # Preparamos los datos para el gráfico, asegurando que haya 7 días
        labels = [(seven_days_ago + timedelta(days=i)).strftime('%b %d') for i in range(7)]
        sales_dict = {s['date'].strftime('%b %d'): float(s['daily_total']) for s in sales_data}
        data = [sales_dict.get(label, 0) for label in labels]

        return JsonResponse({'labels': labels, 'data': data})

    if chart_type == 'top_dishes':
        # Top 5 platos más vendidos
        top_dishes = MenuItem.objects.annotate(
            total_sold=Sum('orderitem__quantity', filter=Q(orderitem__order__status='paid'))
        ).filter(total_sold__gt=0).order_by('-total_sold')[:5]

        labels = [item.name for item in top_dishes]
        data = [item.total_sold for item in top_dishes]
        return JsonResponse({'labels': labels, 'data': data})

    if chart_type == 'sales_by_hour':
        # Ventas de hoy por hora
        sales_data = Order.objects.filter(
            status='paid',
            created_at__date=today
        ).annotate(
            hour=ExtractHour('created_at') # Use the new total_amount field
        ).values('hour').annotate( 
            hourly_total=Sum('total_amount')
        ).order_by('hour')

        labels = [f"{h}:00" for h in range(24)]
        sales_dict = {s['hour']: float(s['hourly_total']) for s in sales_data}
        data = [sales_dict.get(h, 0) for h in range(24)]
        
        return JsonResponse({'labels': labels, 'data': data})

    if chart_type == 'waiter_performance':
        # Rendimiento de garzones (pedidos atendidos hoy)
        waiters = User.objects.filter(groups__name='Garzón')
        performance_data = Order.objects.filter(
            created_at__date=today,
            user__in=waiters
        ).values('user__username').annotate(
            order_count=Count('id'), # Use the new total_amount field
            total_sales=Sum('total_amount', filter=Q(status='paid'))
        ).order_by('-order_count')

        labels = [d['user__username'] for d in performance_data]
        data_orders = [d['order_count'] for d in performance_data]
        data_sales = [float(d['total_sales'] or 0) for d in performance_data]

        return JsonResponse({'labels': labels, 'orders': data_orders, 'sales': data_sales})

    if chart_type == 'sales_by_category':
        # Ventas de hoy por categoría de producto
        from django.db.models import DecimalField
        try:
            category_sales = OrderItem.objects.filter(
                order__status='paid',
                order__created_at__date=today,
                menu_item__isnull=False,
                menu_item__category__isnull=False
            ).values('menu_item__category').annotate(
                total=Sum(F('menu_item__price') * F('quantity'), output_field=DecimalField())
            ).order_by('-total')
            
            labels = []
            data = []
            for item in category_sales:
                category_id = item['menu_item__category']
                if category_id:
                    try:
                        category = Category.objects.get(id=category_id)
                        labels.append(category.name)
                        data.append(float(item['total']))
                    except Category.DoesNotExist:
                        pass
            
            return JsonResponse({'labels': labels, 'data': data})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'labels': [], 'data': []})

    return JsonResponse({'error': 'Invalid chart type'}, status=400)

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Administrador').exists())
def api_admin_dashboard_stats(request):
    """
    API endpoint to get admin dashboard statistics for today.
    Returns: total_today, preparing, ready, completed, total_sales_today
    """
    from django.utils import timezone
    
    try:
        # Usar localtime para obtener la fecha en la zona horaria local
        today = timezone.localtime().date()
        
        # Obtener stats de todas las órdenes (sin filtro de fecha por ahora)
        all_orders = Order.objects.all()
        
        stats = {
            'total_today': all_orders.count(),
            'preparing': all_orders.filter(status='preparing').count(),
            'ready': all_orders.filter(status='ready').count(),
            'completed': all_orders.filter(status='paid').count(),
            'total_sales_today': float(all_orders.filter(status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
        }
        
        return JsonResponse(stats)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Administrador', 'Recepcionista']).exists())
def api_payment_methods_report(request):
    """
    API endpoint to get payment methods statistics with daily and weekly breakdowns.
    Returns monthly data, daily breakdown, and weekly breakdown.
    """
    if request.method != 'GET':
        return JsonResponse({'error': f'Método {request.method} no permitido. Use GET.'}, status=405)
    
    from datetime import datetime, timedelta, date
    from django.db.models import F, Q
    from django.utils import timezone
    
    # Get current month in Santiago timezone
    today = timezone.now().date()
    month_start = today.replace(day=1)
    
    # Get last day of month
    if today.month == 12:
        month_end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    
    # Convert dates to timezone-aware datetimes for filtering
    month_start_dt = timezone.make_aware(datetime.combine(month_start, datetime.min.time()))
    month_end_dt = timezone.make_aware(datetime.combine(month_end, datetime.max.time()))
    
    # Filtrar órdenes pagadas o cargadas a habitación en el mes actual
    orders = Order.objects.filter(
        Q(status='paid') | Q(status='charged_to_room'),
        created_at__gte=month_start_dt,
        created_at__lte=month_end_dt
    )
    
    # === MONTHLY SUMMARY (por método de pago) ===
    payment_stats = orders.values('payment_method').annotate(
        count=Count('id'),
        total_sales=Sum('total_amount'),
        total_tips=Sum('tip_amount')
    ).order_by('-total_sales')
    
    payment_methods = dict(Order.PAYMENT_METHOD_CHOICES)
    monthly_data = []
    grand_total = decimal.Decimal('0.00')
    grand_tips = decimal.Decimal('0.00')
    
    for stat in payment_stats:
        method = stat['payment_method']
        count = stat['count']
        total = stat['total_sales'] or decimal.Decimal('0.00')
        tips = stat['total_tips'] or decimal.Decimal('0.00')
        
        grand_total += total
        grand_tips += tips
        
        monthly_data.append({
            'method': method,
            'method_display': payment_methods.get(method, method),
            'count': count,
            'average': float(total / count) if count > 0 else 0,
            'total': float(total),
            'total_tips': float(tips),
            'percentage': 0
        })
    
    # Calculate percentages
    if grand_total > 0:
        for item in monthly_data:
            item['percentage'] = round((decimal.Decimal(str(item['total'])) / grand_total) * 100, 2)
    
    # === DAILY BREAKDOWN ===
    # Group by date
    daily_dict = {}
    for order in orders:
        order_date = order.created_at.date()
        if order_date not in daily_dict:
            daily_dict[order_date] = {
                'count': 0,
                'total': decimal.Decimal('0.00'),
                'total_tips': decimal.Decimal('0.00')
            }
        daily_dict[order_date]['count'] += 1
        daily_dict[order_date]['total'] += order.total_amount or decimal.Decimal('0.00')
        daily_dict[order_date]['total_tips'] += order.tip_amount or decimal.Decimal('0.00')
    
    daily_data = []
    days_spanish = {
        'Monday': 'Lunes',
        'Tuesday': 'Martes',
        'Wednesday': 'Miércoles',
        'Thursday': 'Jueves',
        'Friday': 'Viernes',
        'Saturday': 'Sábado',
        'Sunday': 'Domingo'
    }
    for date_obj in sorted(daily_dict.keys()):
        stat = daily_dict[date_obj]
        day_name_en = date_obj.strftime('%A')
        day_name_es = days_spanish.get(day_name_en, day_name_en)
        daily_data.append({
            'date': str(date_obj),
            'day_name': day_name_es,
            'count': stat['count'],
            'total': float(stat['total']),
            'total_tips': float(stat['total_tips']),
            'average': float(stat['total'] / stat['count']) if stat['count'] > 0 else 0
        })
    
    # === WEEKLY BREAKDOWN ===
    # Group by ISO week with date range
    from datetime import timedelta
    weekly_stats = {}
    for order in orders:
        year, week_num, weekday = order.created_at.isocalendar()
        week_key = f"W{week_num}"
        
        if week_key not in weekly_stats:
            # Calculate week start (Monday) and end (Sunday)
            order_date = order.created_at.date()
            week_start = order_date - timedelta(days=weekday - 1)  # ISO weekday: Mon=1, Sun=7
            week_end = week_start + timedelta(days=6)
            
            weekly_stats[week_key] = {
                'week_num': week_num,
                'week_start': week_start,
                'week_end': week_end,
                'count': 0,
                'total': decimal.Decimal('0.00'),
                'total_tips': decimal.Decimal('0.00')
            }
        
        weekly_stats[week_key]['count'] += 1
        weekly_stats[week_key]['total'] += order.total_amount or decimal.Decimal('0.00')
        weekly_stats[week_key]['total_tips'] += order.tip_amount or decimal.Decimal('0.00')
    
    weekly_data = []
    months_spanish = {
        'January': 'Ene', 'February': 'Feb', 'March': 'Mar', 'April': 'Abr',
        'May': 'May', 'June': 'Jun', 'July': 'Jul', 'August': 'Ago',
        'September': 'Sep', 'October': 'Oct', 'November': 'Nov', 'December': 'Dic'
    }
    for week_key in sorted(weekly_stats.keys()):
        stat = weekly_stats[week_key]
        # Format as "8-14 Dic" or similar
        start_date = stat['week_start'].strftime('%d')
        month_name_en = stat['week_end'].strftime('%B')
        month_name_es = months_spanish.get(month_name_en, month_name_en)
        end_date = stat['week_end'].strftime('%d')
        week_label = f"{start_date}-{end_date} {month_name_es}"
        
        weekly_data.append({
            'week': week_label,
            'week_num': stat['week_num'],
            'week_start': str(stat['week_start']),
            'week_end': str(stat['week_end']),
            'count': stat['count'],
            'total': float(stat['total']),
            'total_tips': float(stat['total_tips']),
            'average': float(stat['total'] / stat['count']) if stat['count'] > 0 else 0
        })
    
    months_spanish_full = {
        'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo', 'April': 'Abril',
        'May': 'Mayo', 'June': 'Junio', 'July': 'Julio', 'August': 'Agosto',
        'September': 'Septiembre', 'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
    }
    month_name_en = month_start.strftime('%B')
    month_name_es = months_spanish_full.get(month_name_en, month_name_en)
    month_year = f"{month_name_es} {month_start.year}"
    
    return JsonResponse({
        'month': month_year,
        'month_start': str(month_start),
        'month_end': str(month_end),
        'summary': {
            'total_orders': len(orders),
            'grand_total': float(grand_total),
            'grand_tips': float(grand_tips),
            'average_order': float(grand_total / len(orders)) if len(orders) > 0 else 0,
            'average_tip': float(grand_tips / len(orders)) if len(orders) > 0 else 0,
        },
        'monthly': monthly_data,
        'daily': daily_data,
        'weekly': weekly_data
    })

@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrador').exists())
def api_registration_pins(request, pk=None):
    """
    API para gestionar los PINs de registro.
    - GET /api/pins/ -> Devuelve todos los PINs.
    - POST /api/pins/ -> Crea un nuevo PIN para un grupo.
    - DELETE /api/pins/<pk>/ -> Elimina un PIN específico.
    """
    from django.utils import timezone
    import string
    import random

    # --- Manejo de un PIN específico (DELETE) ---
    if pk is not None:
        if request.method == 'DELETE':
            try:
                pin = RegistrationPin.objects.get(pk=pk)
                pin.delete()
                return JsonResponse({'success': True, 'message': 'PIN eliminado'}, status=204) # 204 No Content
            except RegistrationPin.DoesNotExist:
                return JsonResponse({'error': 'PIN no encontrado'}, status=404)
        else:
            # No se permite GET o POST a un PIN específico por ahora
            return JsonResponse({'error': f'Método {request.method} no permitido para esta URL.'}, status=405)

    # --- Manejo de la colección de PINs (GET, POST) ---
    if request.method == 'GET':
        pins = RegistrationPin.objects.select_related('group', 'used_by').order_by('-created_at')
        data = [{
            'id': pin.id,
            'pin': pin.pin,
            'group_name': pin.group.name,
            'is_used': pin.used_by is not None,
            'used_by': pin.used_by.username if pin.used_by else '-',
            'created_at': timezone.localtime(pin.created_at).strftime('%Y-%m-%d %H:%M')
        } for pin in pins]
        return JsonResponse(data, safe=False)

    elif request.method == 'POST':
        data = json.loads(request.body)
        group_id = data.get('group_id')
        if not group_id:
            return JsonResponse({'error': 'El ID del grupo (group_id) es requerido.'}, status=400)
        try:
            group = Group.objects.get(id=group_id)
            # Generar un PIN aleatorio y único
            while True:
                pin_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))
                if not RegistrationPin.objects.filter(pin=pin_code).exists():
                    break
            pin = RegistrationPin.objects.create(pin=pin_code, group=group)
            return JsonResponse({'success': True, 'pin': pin.pin, 'group': group.name}, status=201)
        except Group.DoesNotExist:
            return JsonResponse({'error': 'El grupo seleccionado no existe.'}, status=400)
    
    # Si llega aquí, es un método no manejado para la URL base
    return JsonResponse({'error': f'Método {request.method} no permitido.'}, status=405)

@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name__in=['Administrador', 'Recepcionista']).exists())
def export_orders_excel(request):
    """
    Exports a filtered list of orders to an Excel file (.xlsx).
    """
    from django.http import HttpResponse
    from django.utils import timezone

    orders = Order.objects.prefetch_related('orderitem_set__menu_item').all()

    # Filtering (same logic as api_orders_report)
    search_query = request.GET.get('search', '')
    if search_query:
        orders = orders.filter(Q(id__icontains=search_query) | Q(client_identifier__icontains=search_query) | Q(room_number__icontains=search_query))

    status_query = request.GET.get('status', '')
    if status_query:
        orders = orders.filter(status=status_query)

    date_from_query = request.GET.get('date_from', '')
    if date_from_query:
        orders = orders.filter(created_at__date__gte=date_from_query)

    date_to_query = request.GET.get('date_to', '')
    if date_to_query:
        orders = orders.filter(created_at__date__lte=date_to_query)

    orders = orders.order_by('-created_at')

    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Pedidos"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    total_font = Font(bold=True)
    currency_format = '"$"#,##0'

    # Write headers
    headers = ['ID Order', 'Cliente/Habitación', 'Estado', 'Método de Pago', 'Fecha y Hora', 'Total']
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Write data rows
    grand_total = 0
    row_num = 2
    grand_total = decimal.Decimal('0.00')
    for order in orders:
        final_total = order.total_amount # Use the new field
        if final_total:
            grand_total += final_total
        local_time = timezone.localtime(order.created_at).strftime('%Y-%m-%d %H:%M')
        
        ws.cell(row=row_num, column=1, value=order.id)
        ws.cell(row=row_num, column=2, value=get_order_identifier(order))
        ws.cell(row=row_num, column=3, value=order.get_status_display())
        ws.cell(row=row_num, column=4, value=order.get_payment_method_display())
        ws.cell(row=row_num, column=5, value=local_time)
        total_cell = ws.cell(row=row_num, column=6, value=final_total)
        total_cell.number_format = currency_format
        row_num += 1

    # Write total row
    total_label_cell = ws.cell(row=row_num + 1, column=5, value="Total Vendido:")
    total_label_cell.font = total_font
    total_label_cell.alignment = Alignment(horizontal='right')

    grand_total_cell = ws.cell(row=row_num + 1, column=6, value=grand_total)
    grand_total_cell.font = total_font
    grand_total_cell.number_format = currency_format
    grand_total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow fill

    # Adjust column widths
    for col_num, header_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    # Create the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_pedidos_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'
    wb.save(response)

    return response

@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administrador').exists())
@csrf_exempt
def api_users(request, pk=None):
    """
    API endpoint to manage users.
    - GET /api/users/ -> List all users.
    - PUT /api/users/<pk>/ -> Update a user's role.
    - DELETE /api/users/<pk>/ -> Delete a user.
    """
    # --- Manejo de un usuario específico (PUT, DELETE) ---
    if pk:
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

        # Evitar que un administrador se modifique o elimine a sí mismo
        if user == request.user:
            return JsonResponse({'error': 'No puedes modificar o eliminar tu propia cuenta.'}, status=403)

        if request.method == 'PUT':
            data = json.loads(request.body)
            group_id = data.get('group_id')
            try:
                group = Group.objects.get(id=group_id)
                user.groups.set([group]) # set() reemplaza todos los grupos existentes
                return JsonResponse({'success': True, 'message': 'Rol de usuario actualizado.'})
            except Group.DoesNotExist:
                return JsonResponse({'error': 'El rol especificado no existe.'}, status=400)

        elif request.method == 'DELETE':
            user.delete()
            return JsonResponse({'success': True, 'message': 'Usuario eliminado.'}, status=204)

    # --- Manejo de la lista de usuarios (GET) ---
    elif request.method == 'GET':
        users = User.objects.prefetch_related('groups').all().order_by('username')
        data = [{
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'group_id': user.groups.first().id if user.groups.exists() else None,
            'group_name': user.groups.first().name if user.groups.exists() else 'N/A',
            'last_login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Nunca'
        } for user in users]
        return JsonResponse(data, safe=False)

    # Método no permitido para la URL solicitada
    return JsonResponse({'error': f'Método {request.method} no permitido.'}, status=405)

def serve_media_file(request, file_path):
    """
    Serve media files (images, documents, etc.) from the media directory.
    This is needed for production environments where DEBUG=False.
    """
    import os
    from django.http import FileResponse
    
    # Secure the file path to prevent directory traversal attacks
    media_root = settings.MEDIA_ROOT
    requested_file = os.path.join(media_root, file_path)
    
    # Normalize paths and ensure file is within MEDIA_ROOT
    requested_file = os.path.normpath(requested_file)
    media_root = os.path.normpath(media_root)
    
    if not requested_file.startswith(media_root):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if not os.path.exists(requested_file) or not os.path.isfile(requested_file):
        return JsonResponse({'error': 'File not found'}, status=404)
    
    try:
        return FileResponse(open(requested_file, 'rb'), content_type='image/jpeg')
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def api_categories(request):
    """
    API para gestionar categorías.
    - GET /api/categories/ -> Devuelve todas las categorías (público).
    - POST /api/categories/ -> Crea una nueva categoría (requiere admin).
    - DELETE /api/categories/<pk>/ -> Elimina una categoría (requiere admin).
    """
    if request.method == 'GET':
        # GET es público - no requiere autenticación
        categories = Category.objects.all().order_by('name')
        data = [{
            'id': cat.id,
            'name': cat.name,
            'description': cat.description or '',
            'created_at': cat.created_at.isoformat()
        } for cat in categories]
        return JsonResponse(data, safe=False)
    
    elif request.method == 'POST':
        # POST requiere que sea admin
        if not (request.user.is_authenticated and (request.user.is_superuser or request.user.groups.filter(name='Administrador').exists())):
            return JsonResponse({'error': 'No tienes permiso para crear categorías.'}, status=403)
        
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            
            if not name:
                return JsonResponse({'error': 'El nombre de la categoría es requerido.'}, status=400)
            
            # Verificar si la categoría ya existe (case-insensitive)
            if Category.objects.filter(name__iexact=name).exists():
                return JsonResponse({'error': 'Esta categoría ya existe.'}, status=400)
            
            category = Category.objects.create(
                name=name,
                description=data.get('description', '')
            )
            
            return JsonResponse({
                'id': category.id,
                'name': category.name,
                'description': category.description or '',
                'created_at': category.created_at.isoformat()
            }, status=201)
        
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido.'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'Error al crear categoría: {str(e)}'}, status=500)
    
    elif request.method == 'DELETE':
        # DELETE requiere que sea admin
        if not (request.user.is_authenticated and (request.user.is_superuser or request.user.groups.filter(name='Administrador').exists())):
            return JsonResponse({'error': 'No tienes permiso para eliminar categorías.'}, status=403)
        
        # Eliminar una categoría específica
        pk = request.GET.get('id')
        if pk:
            try:
                category = Category.objects.get(pk=pk)
                # Verificar si hay items usando esta categoría
                if MenuItem.objects.filter(category=category.name).exists():
                    return JsonResponse({'error': 'No se puede eliminar una categoría que está en uso.'}, status=400)
                category.delete()
                return JsonResponse({'success': True, 'message': 'Categoría eliminada.'}, status=204)
            except Category.DoesNotExist:
                return JsonResponse({'error': 'Categoría no encontrada.'}, status=404)
        else:
            return JsonResponse({'error': 'ID de categoría requerido.'}, status=400)
    
    return JsonResponse({'error': f'Método {request.method} no permitido.'}, status=405)


@csrf_exempt
def api_categories_check(request):
    """
    Verifica si una categoría existe (case-insensitive) y detecta similares.
    GET /api/categories/check/?name=Entradas
    
    Retorna:
    {
        'name': 'Entradas',
        'exists': bool,
        'similar': [
            {'id': 1, 'name': 'entrada', 'similarity': 0.95}
        ]
    }
    """
    if request.method == 'GET':
        from difflib import SequenceMatcher
        
        name = request.GET.get('name', '').strip()
        
        if not name:
            return JsonResponse({'error': 'El nombre de la categoría es requerido.'}, status=400)
        
        # Verificar si existe exactamente (case-insensitive)
        exists = Category.objects.filter(name__iexact=name).exists()
        
        # Encontrar categorías similares
        all_categories = Category.objects.all()
        similar = []
        
        for cat in all_categories:
            # Calcular similitud usando SequenceMatcher
            similarity = SequenceMatcher(None, name.lower(), cat.name.lower()).ratio()
            
            # Si la similitud es > 70% y NO es exactamente igual (ese caso ya lo cubrimos con exists)
            if similarity > 0.7 and not cat.name.lower() == name.lower():
                similar.append({
                    'id': cat.id,
                    'name': cat.name,
                    'similarity': round(similarity * 100, 1)  # Porcentaje de similitud
                })
        
        # Ordenar por similitud (descendente)
        similar.sort(key=lambda x: x['similarity'], reverse=True)
        
        return JsonResponse({
            'name': name,
            'exists': exists,
            'similar': similar
        })
    
    return JsonResponse({'error': f'Método {request.method} no permitido.'}, status=405)


@csrf_exempt
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Recepcionista').exists())
def api_process_payment(request, pk):
    """
    API endpoint to process payment for an order.
    PUT: Accepts payment_method, tip_amount, and optional payment_reference.
    Updates order status to 'paid' and records payment details.
    """
    from django.utils import timezone
    
    if request.method != 'PUT':
        return JsonResponse({'error': f'Método {request.method} no permitido. Use PUT.'}, status=405)
    
    try:
        order = Order.objects.get(pk=pk)
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Orden no encontrada'}, status=404)
    
    try:
        data = json.loads(request.body)
        payment_method = data.get('payment_method', 'cash')
        tip_amount = decimal.Decimal(str(data.get('tip_amount', 0)))
        payment_reference = data.get('payment_reference', '')
        
        # Validar método de pago válido
        valid_methods = dict(Order.PAYMENT_METHOD_CHOICES).keys()
        if payment_method not in valid_methods:
            return JsonResponse({'error': 'Método de pago inválido'}, status=400)
        
        # Validar que la propina no sea negativa
        if tip_amount < 0:
            return JsonResponse({'error': 'La propina no puede ser negativa'}, status=400)
        
        # Actualizar orden
        with transaction.atomic():
            order.payment_method = payment_method
            order.paid_at = timezone.now()
            order.status = 'paid'
            order.tip_amount = tip_amount
            if payment_reference:
                order.payment_reference = payment_reference
            # El total ya incluye los items, solo sumamos la propina
            order.total_amount = sum(
                item.menu_item.price * item.quantity 
                for item in order.orderitem_set.all()
            ) + tip_amount
            order.save()
        
        return JsonResponse({
            'success': True,
            'order_id': order.id,
            'status': order.get_status_display(),
            'payment_method': order.get_payment_method_display(),
            'total_amount': float(order.total_amount),
            'tip_amount': float(order.tip_amount),
            'paid_at': order.paid_at.isoformat() if order.paid_at else None
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido en el cuerpo de la solicitud'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error al procesar pago: {str(e)}'}, status=500)


# ============ APIS PARA ROOMBILL ============

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Recepcionista').exists())
@csrf_exempt
def api_get_unpaid_orders_by_room(request):
    """
    GET: Retorna los pedidos sin pagar agrupados por habitación y dentro de cada habitación por cliente
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        from django.db.models import Sum
        
        # Obtener solo pedidos servidos que no estén pagados
        unpaid_orders = Order.objects.filter(
            status__in=['served', 'charged_to_room']
        ).select_related('user').order_by('room_number', 'client_identifier', '-created_at')
        
        # Agrupar por habitación y dentro de cada habitación por cliente
        rooms = {}
        for order in unpaid_orders:
            room = order.room_number or 'Sin Habitación'
            client = order.client_identifier or 'Sin nombre'
            
            # Crear estructura de habitación si no existe
            if room not in rooms:
                rooms[room] = {
                    'room_number': room,
                    'clients': {},
                    'total': 0
                }
            
            # Crear estructura de cliente dentro de la habitación si no existe
            if client not in rooms[room]['clients']:
                rooms[room]['clients'][client] = {
                    'name': client,
                    'orders': [],
                    'total': 0
                }
            
            # Agregar pedido al cliente
            rooms[room]['clients'][client]['orders'].append({
                'id': order.id,
                'created_at': order.created_at.isoformat(),
                'status': order.status,
                'status_display': order.get_status_display(),
                'items': [
                    {
                        'name': item.menu_item.name,
                        'quantity': item.quantity,
                        'price': float(item.menu_item.price),
                        'subtotal': float(item.menu_item.price * item.quantity)
                    }
                    for item in order.orderitem_set.all()
                ],
                'total': float(order.total_amount),
                'tip': float(order.tip_amount)
            })
            
            # Actualizar totales
            rooms[room]['clients'][client]['total'] += float(order.total_amount)
            rooms[room]['total'] += float(order.total_amount)
        
        # Convertir diccionarios de clientes a listas
        for room in rooms.values():
            room['clients'] = list(room['clients'].values())
        
        return JsonResponse({'rooms': list(rooms.values())}, safe=False)
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Recepcionista').exists())
@csrf_exempt
def api_create_roombill(request):
    """
    POST: Crea una nueva factura de habitación con los pedidos seleccionados
    Esperado: {
        "room_number": "101",
        "guest_name": "John Doe",
        "order_ids": [1, 2, 3],
        "tip_amount": 10.00
    }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        from django.utils import timezone
        data = json.loads(request.body)
        
        room_number = data.get('room_number', '').strip()
        guest_name = data.get('guest_name', '').strip()
        order_ids = data.get('order_ids', [])
        tip_amount = decimal.Decimal(str(data.get('tip_amount', '0.00')))
        
        if not room_number:
            return JsonResponse({'error': 'Habitación requerida'}, status=400)
        
        if not order_ids:
            return JsonResponse({'error': 'Debe seleccionar al menos un pedido'}, status=400)
        
        # Validar que todos los pedidos existan y sean de la misma habitación
        orders = Order.objects.filter(id__in=order_ids)
        
        if len(orders) != len(order_ids):
            return JsonResponse({'error': 'Algunos pedidos no existen'}, status=400)
        
        # Validar que todos sean de la misma habitación
        for order in orders:
            if order.room_number != room_number:
                return JsonResponse({'error': f'El pedido {order.id} no es de la habitación {room_number}'}, status=400)
        
        # Crear la factura
        from .models import RoomBill
        bill = RoomBill.objects.create(
            room_number=room_number,
            guest_name=guest_name,
            tip_amount=tip_amount,
            created_by=request.user,
            status='draft'
        )
        
        # Agregar los pedidos
        bill.orders.set(orders)
        
        # Calcular el total
        total = bill.calculate_total() + tip_amount
        bill.total_amount = total
        bill.save()
        
        return JsonResponse({
            'success': True,
            'bill_id': bill.id,
            'message': f'Factura creada: {bill.id}'
        }, status=201)
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Recepcionista').exists())
@csrf_exempt
def api_get_roombills(request):
    """
    GET: Retorna las facturas de habitación filtradas por estado
    ?status=draft|confirmed|paid|all
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        from .models import RoomBill
        
        status_filter = request.GET.get('status', 'all')
        
        bills = RoomBill.objects.prefetch_related('orders').all()
        
        if status_filter != 'all':
            bills = bills.filter(status=status_filter)
        
        bills = bills.order_by('-created_at')
        
        data = []
        for bill in bills:
            data.append({
                'id': bill.id,
                'room_number': bill.room_number,
                'guest_name': bill.guest_name,
                'status': bill.status,
                'status_display': bill.get_status_display(),
                'status_class': bill.status_class,
                'total': float(bill.total_amount),
                'tip': float(bill.tip_amount),
                'payment_method': bill.payment_method,
                'payment_method_display': bill.get_payment_method_display(),
                'order_count': bill.orders.count(),
                'created_at': bill.created_at.isoformat(),
                'paid_at': bill.paid_at.isoformat() if bill.paid_at else None,
            })
        
        return JsonResponse({'bills': data}, safe=False)
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Recepcionista').exists())
@csrf_exempt
def api_roombill_detail(request, bill_id):
    """
    GET: Obtiene los detalles de una factura
    POST: Actualiza el estado o método de pago
    """
    try:
        from .models import RoomBill
        bill = RoomBill.objects.prefetch_related('orders__orderitem_set__menu_item').get(id=bill_id)
    except RoomBill.DoesNotExist:
        return JsonResponse({'error': 'Factura no encontrada'}, status=404)
    
    if request.method == 'GET':
        orders_data = []
        for order in bill.orders.all():
            orders_data.append({
                'id': order.id,
                'created_at': order.created_at.isoformat(),
                'items': [
                    {
                        'name': item.menu_item.name,
                        'quantity': item.quantity,
                        'price': float(item.menu_item.price),
                        'subtotal': float(item.menu_item.price * item.quantity)
                    }
                    for item in order.orderitem_set.all()
                ],
                'total': float(order.total_amount)
            })
        
        return JsonResponse({
            'id': bill.id,
            'room_number': bill.room_number,
            'guest_name': bill.guest_name,
            'status': bill.status,
            'status_display': bill.get_status_display(),
            'total': float(bill.total_amount),
            'tip': float(bill.tip_amount),
            'payment_method': bill.payment_method,
            'payment_method_display': bill.get_payment_method_display(),
            'orders': orders_data,
            'created_at': bill.created_at.isoformat()
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Actualizar estado si se proporciona
            if 'status' in data:
                new_status = data.get('status')
                if new_status not in ['draft', 'confirmed', 'paid', 'cancelled']:
                    return JsonResponse({'error': 'Estado inválido'}, status=400)
                bill.status = new_status
            
            # Actualizar método de pago si se proporciona
            if 'payment_method' in data:
                payment_method = data.get('payment_method')
                if payment_method not in ['cash', 'card', 'transfer', 'check', 'mixed']:
                    return JsonResponse({'error': 'Método de pago inválido'}, status=400)
                bill.payment_method = payment_method
            
            # Si se marca como pagada, actualizar paid_at
            if bill.status == 'paid' and not bill.paid_at:
                from django.utils import timezone
                bill.paid_at = timezone.now()
                
                # Actualizar el estado de los pedidos a 'paid'
                for order in bill.orders.all():
                    order.status = 'paid'
                    order.paid_at = bill.paid_at
                    order.save()
            
            bill.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Factura actualizada',
                'status': bill.get_status_display()
            })
        
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)


@login_required
def export_roombills_excel(request):
    """
    Exporta un reporte de facturas de habitación a un archivo Excel.
    """
    from django.http import HttpResponse
    from django.utils import timezone

    # Obtener todas las facturas
    bills = RoomBill.objects.prefetch_related('orders__orderitem_set').all()

    # Filtrado opcional por estado (puede ser múltiple separado por comas)
    status_query = request.GET.get('status', '')
    if status_query:
        if status_query != 'all':
            statuses = status_query.split(',')
            bills = bills.filter(status__in=statuses)

    # Filtrado opcional por fecha
    date_from_query = request.GET.get('date_from', '')
    if date_from_query:
        bills = bills.filter(created_at__date__gte=date_from_query)

    date_to_query = request.GET.get('date_to', '')
    if date_to_query:
        bills = bills.filter(created_at__date__lte=date_to_query)

    # Filtrado opcional por habitación
    room_query = request.GET.get('room', '')
    if room_query:
        bills = bills.filter(room_number__icontains=room_query)

    bills = bills.order_by('-created_at')

    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Facturas"

    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6F4E37", end_color="6F4E37", fill_type="solid")  # Color café
    total_font = Font(bold=True)
    currency_format = '"$"#,##0.00'
    date_format = 'yyyy-mm-dd hh:mm'

    # Escribir encabezados
    headers = ['ID Factura', 'Habitación', 'Estado', 'Subtotal', 'Propina', 'Total', 'Método de Pago', 'Fecha de Creación', 'Fecha de Pago']
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Escribir datos
    grand_total = decimal.Decimal('0.00')
    grand_subtotal = decimal.Decimal('0.00')
    grand_tip = decimal.Decimal('0.00')
    row_num = 2

    for bill in bills:
        subtotal = bill.total_amount - bill.tip_amount
        grand_total += bill.total_amount
        grand_subtotal += subtotal
        grand_tip += bill.tip_amount

        local_time = timezone.localtime(bill.created_at).strftime('%Y-%m-%d %H:%M')
        paid_time = timezone.localtime(bill.paid_at).strftime('%Y-%m-%d %H:%M') if bill.paid_at else '-'

        ws.cell(row=row_num, column=1, value=bill.id)
        ws.cell(row=row_num, column=2, value=bill.room_number)
        ws.cell(row=row_num, column=3, value=bill.get_status_display())

        subtotal_cell = ws.cell(row=row_num, column=4, value=float(subtotal))
        subtotal_cell.number_format = currency_format

        tip_cell = ws.cell(row=row_num, column=5, value=float(bill.tip_amount))
        tip_cell.number_format = currency_format

        total_cell = ws.cell(row=row_num, column=6, value=float(bill.total_amount))
        total_cell.number_format = currency_format

        ws.cell(row=row_num, column=7, value=bill.get_payment_method_display() if bill.payment_method else '-')
        ws.cell(row=row_num, column=8, value=local_time)
        ws.cell(row=row_num, column=9, value=paid_time)

        row_num += 1

    # Escribir fila de totales
    # Fila de Subtotal
    subtotal_label_cell = ws.cell(row=row_num + 1, column=3, value="Subtotal:")
    subtotal_label_cell.font = total_font
    subtotal_label_cell.alignment = Alignment(horizontal='right')

    subtotal_total_cell = ws.cell(row=row_num + 1, column=4, value=float(grand_subtotal))
    subtotal_total_cell.font = total_font
    subtotal_total_cell.number_format = currency_format
    subtotal_total_cell.fill = PatternFill(start_color="FCD34D", end_color="FCD34D", fill_type="solid")

    # Fila de Propina
    tip_label_cell = ws.cell(row=row_num + 2, column=3, value="Propina:")
    tip_label_cell.font = total_font
    tip_label_cell.alignment = Alignment(horizontal='right')

    tip_total_cell = ws.cell(row=row_num + 2, column=5, value=float(grand_tip))
    tip_total_cell.font = total_font
    tip_total_cell.number_format = currency_format
    tip_total_cell.fill = PatternFill(start_color="FCD34D", end_color="FCD34D", fill_type="solid")

    # Fila de Total con Propina
    final_total_label_cell = ws.cell(row=row_num + 3, column=3, value="TOTAL CON PROPINA:")
    final_total_label_cell.font = Font(bold=True, size=12)
    final_total_label_cell.alignment = Alignment(horizontal='right')

    grand_total_cell = ws.cell(row=row_num + 3, column=6, value=float(grand_total))
    grand_total_cell.font = Font(bold=True, size=12)
    grand_total_cell.number_format = currency_format
    grand_total_cell.fill = PatternFill(start_color="FCD34D", end_color="FCD34D", fill_type="solid")

    # Ajustar ancho de columnas
    for col_num, header_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 18

    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_facturas_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'
    wb.save(response)

    return response